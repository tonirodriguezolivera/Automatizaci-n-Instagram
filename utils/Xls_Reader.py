import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, NamedStyle
from openpyxl.utils import get_column_letter

class XlsReader:
    """
    Esta clase proporciona métodos para leer y escribir datos desde y hacia un archivo Excel (XLSX).
    Autor: Kevin Arnold
    """
    
    def __init__(self, path):
        """
        Constructor de la clase XlsReader.

        Args:
            path (str): La ruta del archivo Excel.
        """
        self.path = path
        self.workbook = None
        self.sheet = None
        try:
            self.workbook = load_workbook(path)
            self.sheet = self.workbook[self.workbook.sheetnames[0]]
        except Exception as e:
            print(f"Error: {str(e)}")

    def get_row_count(self, sheet_name):
        """
        Obtiene el número de filas en una hoja de Excel.

        Args:
            sheet_name (str): El nombre de la hoja de Excel.

        Returns:
            int: El número de filas en la hoja especificada.
        """
        try:
            if sheet_name not in self.workbook.sheetnames:
                return 0
            sheet = self.workbook[sheet_name]
            return sheet.max_row
        except Exception as e:
            print(f"Error: {str(e)}")
            return 0

    def get_cell_data(self, sheet_name, col_name=None, row_num=None, col_num=None):
        """
        Obtiene el valor de una celda en una hoja de Excel.

        Args:
            sheet_name (str): El nombre de la hoja de Excel.
            col_name (str, optional): El nombre de la columna.
            row_num (int, optional): El número de fila.
            col_num (int, optional): El número de columna.

        Returns:
            str: El valor de la celda en formato de cadena.
        """
        try:
            if row_num is not None and row_num <= 0:
                return ""

            if sheet_name not in self.workbook.sheetnames:
                return ""

            sheet = self.workbook[sheet_name]
            
            if col_name is not None:
                # Find column index by name
                header_row = sheet[1]
                col_num = None
                for i, cell in enumerate(header_row, 1):
                    if cell.value and str(cell.value).strip() == col_name.strip():
                        col_num = i
                        break
                if col_num is None:
                    return ""

            # Get cell value
            cell = sheet.cell(row=row_num, column=col_num)
            
            if cell.value is None:
                return ""

            if isinstance(cell.value, (int, float)):
                if cell.value == int(cell.value):
                    return str(int(cell.value))
                return str(cell.value)
            return str(cell.value)

        except Exception as e:
            print(f"Error: {str(e)}")
            return f"row {row_num} or column {col_name or col_num} does not exist in xls"

    def get_cell_data_by_row_name(self, sheet_name, col_name, row_name):
        """
        Obtiene el valor de una celda por nombre de fila y columna.

        Args:
            sheet_name (str): El nombre de la hoja de Excel.
            col_name (str): El nombre de la columna.
            row_name (str): El nombre/identificador de la fila.

        Returns:
            str: El valor de la celda en formato de cadena.
        """
        try:
            if not row_name or not row_name.strip():
                return ""

            if sheet_name not in self.workbook.sheetnames:
                return ""

            sheet = self.workbook[sheet_name]
            
            # Find column index
            header_row = sheet[1]
            col_num = None
            for i, cell in enumerate(header_row, 1):
                if cell.value and str(cell.value).strip() == col_name.strip():
                    col_num = i
                    break
            if col_num is None:
                return ""

            # Find row index by name
            row_index = self._find_row_index_by_name(sheet, row_name)
            if row_index == -1:
                return ""

            cell = sheet.cell(row=row_index, column=col_num)
            
            if cell.value is None:
                return ""

            return str(cell.value)

        except Exception as e:
            print(f"Error: {str(e)}")
            return f"Error: {str(e)}"

    def _find_row_index_by_name(self, sheet, target_name):
        """
        Encuentra el índice de la fila por nombre/identificador en la primera columna.

        Args:
            sheet: La hoja de Excel.
            target_name (str): El nombre/identificador a buscar.

        Returns:
            int: El índice de la fila o -1 si no se encuentra.
        """
        for i in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=i, column=1)
            if cell.value and str(cell.value).strip() == target_name.strip():
                return i
        return -1

    def set_cell_data(self, sheet_name, col_name, row_num, data):
        """
        Establece el valor de una celda en una hoja de Excel.

        Args:
            sheet_name (str): El nombre de la hoja de Excel.
            col_name (str): El nombre de la columna.
            row_num (int): El número de fila.
            data (str): El nuevo valor de la celda.

        Returns:
            bool: True si la operación tiene éxito, False en caso contrario.
        """
        try:
            if row_num <= 0:
                return False

            if sheet_name not in self.workbook.sheetnames:
                return False

            sheet = self.workbook[sheet_name]
            
            # Find column index
            header_row = sheet[1]
            col_num = None
            for i, cell in enumerate(header_row, 1):
                if cell.value and str(cell.value).strip() == col_name.strip():
                    col_num = i
                    break
            if col_num is None:
                return False

            # Create row if it doesn't exist
            if sheet.max_row < row_num:
                sheet.cell(row=row_num, column=col_num).value = data
            else:
                sheet.cell(row=row_num, column=col_num).value = data

            self.workbook.save(self.path)
            return True

        except Exception as e:
            print(f"Error: {str(e)}")
            return False

    def add_sheet(self, sheet_name):
        """
        Agrega una nueva hoja de Excel.

        Args:
            sheet_name (str): El nombre de la nueva hoja.

        Returns:
            bool: True si la operación tiene éxito, False en caso contrario.
        """
        try:
            self.workbook.create_sheet(sheet_name)
            self.workbook.save(self.path)
            return True
        except Exception as e:
            print(f"Error: {str(e)}")
            return False

    def remove_sheet(self, sheet_name):
        """
        Elimina una hoja de Excel.

        Args:
            sheet_name (str): El nombre de la hoja a eliminar.

        Returns:
            bool: True si la operación tiene éxito, False en caso contrario.
        """
        try:
            if sheet_name not in self.workbook.sheetnames:
                return False
            del self.workbook[sheet_name]
            self.workbook.save(self.path)
            return True
        except Exception as e:
            print(f"Error: {str(e)}")
            return False

    def add_column(self, sheet_name, col_name):
        """
        Agrega una nueva columna a una hoja de Excel.

        Args:
            sheet_name (str): El nombre de la hoja de Excel.
            col_name (str): El nombre de la nueva columna.

        Returns:
            bool: True si la operación tiene éxito, False en caso contrario.
        """
        try:
            if sheet_name not in self.workbook.sheetnames:
                return False

            sheet = self.workbook[sheet_name]
            
            # Create style for header
            style = NamedStyle(name=f"header_{col_name}")
            style.fill = PatternFill(start_color="A0A0A0", end_color="A0A0A0", fill_type="solid")

            # Add column header
            last_column = sheet.max_column + 1 if sheet.max_column else 1
            sheet.cell(row=1, column=last_column).value = col_name
            sheet.cell(row=1, column=last_column).style = style

            # Auto-size column
            sheet.column_dimensions[get_column_letter(last_column)].width = len(col_name) + 2

            self.workbook.save(self.path)
            return True

        except Exception as e:
            print(f"Error: {str(e)}")
            return False

    def remove_column(self, sheet_name, col_num):
        """
        Elimina una columna de una hoja de Excel.

        Args:
            sheet_name (str): El nombre de la hoja de Excel.
            col_num (int): El número de columna a eliminar.

        Returns:
            bool: True si la operación tiene éxito, False en caso contrario.
        """
        try:
            if not self.is_sheet_exist(sheet_name):
                return False

            sheet = self.workbook[sheet_name]
            
            for row in range(1, sheet.max_row + 1):
                sheet.cell(row=row, column=col_num).value = None

            self.workbook.save(self.path)
            return True

        except Exception as e:
            print(f"Error: {str(e)}")
            return False

    def is_sheet_exist(self, sheet_name):
        """
        Verifica si una hoja de Excel existe.

        Args:
            sheet_name (str): El nombre de la hoja a verificar.

        Returns:
            bool: True si la hoja existe, False en caso contrario.
        """
        return sheet_name in self.workbook.sheetnames or sheet_name.upper() in self.workbook.sheetnames

    def get_column_count(self, sheet_name):
        """
        Obtiene el número de columnas en una hoja de Excel.

        Args:
            sheet_name (str): El nombre de la hoja de Excel.

        Returns:
            int: El número de columnas en la hoja especificada.
        """
        if not self.is_sheet_exist(sheet_name):
            return -1

        sheet = self.workbook[sheet_name]
        return sheet.max_column

    def get_case_data_for_country(self, sheet_name, workflow, case_name, country):
        """
        Obtiene el valor de una celda específica según el flujo de trabajo, el nombre del caso y el país.

        Args:
            sheet_name (str): El nombre de la hoja de Excel.
            workflow (str): El número del flujo de trabajo.
            case_name (str): El nombre del caso en el flujo de trabajo.
            country (str): El nombre del país (columna).

        Returns:
            str: El valor de la celda correspondiente o una cadena vacía si no se encuentra.
        """
        try:
            if sheet_name not in self.workbook.sheetnames:
                print(f"Sheet not found: {sheet_name}")
                return ""

            sheet = self.workbook[sheet_name]
            case_column = 3  # Column C (CASO)
            
            for i in range(2, sheet.max_row + 1):
                cell_value = sheet.cell(row=i, column=case_column).value
                if cell_value and str(cell_value).strip() == case_name.strip():
                    workflow_value = self.get_cell_data(sheet_name, "Workflow #", row_num=i)
                    if workflow_value == workflow:
                        data = self.get_cell_data(sheet_name, country, row_num=i)
                        # Note: ScenarioManager.setNameWorkflow is not implemented in Python
                        return data

            return ""

        except Exception as e:
            print(f"Error: {str(e)}")
            return ""

    def get_data_for_workflow_and_country(self, sheet_name, workflow, country):
        """
        Obtiene el valor de una celda según el flujo de trabajo y el país.

        Args:
            sheet_name (str): El nombre de la hoja de Excel.
            workflow (str): El número del flujo de trabajo.
            country (str): El nombre del país (columna).

        Returns:
            str: El valor de la celda correspondiente o una cadena vacía si no se encuentra.
        """
        try:
            if sheet_name not in self.workbook.sheetnames:
                print(f"Sheet not found: {sheet_name}")
                return ""

            sheet = self.workbook[sheet_name]
            
            for i in range(2, sheet.max_row + 1):
                workflow_value = self.get_cell_data(sheet_name, "Workflow #", row_num=i)
                if workflow_value == workflow:
                    data = self.get_cell_data(sheet_name, country, row_num=i)
                    # Note: ScenarioManager.setNameWorkflow is not implemented in Python
                    return data

            return ""

        except Exception as e:
            print(f"Error: {str(e)}")
            return ""